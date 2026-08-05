"""อ่านรายชื่อผู้ป่วยจากไฟล์ Excel (คอลัมน์ an — หรือ hn ถ้ามี DB ให้แปลงเป็น an ได้)"""
import io

from openpyxl import load_workbook


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _pick_sheet(wb):
    """ใช้ชีตแรกที่มองเห็นได้ — ชีตที่ถูกซ่อนมักเป็นข้อมูลเก่า/ตารางช่วยคำนวณ"""
    for ws in wb.worksheets:
        if getattr(ws, "sheet_state", "visible") == "visible":
            return ws
    return wb.worksheets[0]


def parse_patients(file_bytes: bytes) -> dict:
    """คืน {"sheet": ชื่อชีต, "columns": [...], "rows": [{an,hn,ptname,ward,...}]}
    ต้องมีหัวคอลัมน์ an หรือ hn (ตัวเล็ก/ใหญ่ได้) คอลัมน์อื่นเก็บติดมาด้วยถ้าชื่อตรง"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = _pick_sheet(wb)
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return {"sheet": ws.title, "columns": [], "rows": []}

        cols = [_cell_str(h).lower() for h in header]
        known = {"an", "hn", "ptname", "ward", "ward_name"}
        idx = {}
        for i, name in enumerate(cols):
            if name in known and name not in idx:
                idx[name] = i  # หัวคอลัมน์ซ้ำ ให้ยึดอันแรก
        if "an" not in idx and "hn" not in idx:
            raise ValueError("ไฟล์ Excel ต้องมีหัวคอลัมน์ชื่อ an หรือ hn ในชีตแรกที่มองเห็น")

        out = []
        seen = set()
        for row in rows_iter:
            rec = {k: _cell_str(row[i]) if i < len(row) else "" for k, i in idx.items()}
            an = rec.get("an", "")
            hn = rec.get("hn", "")
            if not an and not hn:
                continue
            key = an or ("hn:" + hn)
            if key in seen:
                continue
            seen.add(key)
            out.append(
                {
                    "an": an,
                    "hn": hn,
                    "ward": rec.get("ward", ""),
                    "ward_name": rec.get("ward_name", ""),
                    "ptname": rec.get("ptname", ""),
                    "regdate": "",
                    "los_days": "",
                }
            )
        return {"sheet": ws.title, "columns": cols, "rows": out}
    finally:
        wb.close()
